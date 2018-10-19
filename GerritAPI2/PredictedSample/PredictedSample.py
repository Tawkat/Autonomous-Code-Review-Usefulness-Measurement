
'''
import sys
reload(sys)
sys.setdefaultencoding('utf8')

import sqlite3
import matplotlib.pyplot as plt
import numpy as np
import operator


connection=sqlite3.connect('Gerrit.db')
connection.text_factory=str
cursor=connection.cursor()

cursor.execute("SELECT DISTINCT R1.reviewData,R1.usefulness from ReviewTable R1 JOIN ChangeTable C ON R1.change_id=C.change_id "
               "WHERE C.project_name='iotivity'")
data=cursor.fetchall()
'''

from Classification.ReviewModifier.ReviewModifier import ReviewModifier
from Classification.Classifier import Classifier
from openpyxl import Workbook,load_workbook

book = load_workbook('PredictedSample - Raw.xlsx')

modified = Workbook()
clf=Classifier()


#for i,d in enumerate(data,2):
sheet = book.active
sheetWrite=modified.active
#book.template=False

matched=0
total=0
for i in range(2,302):
    id=sheet['A'+str(900+i)]
    review=sheet['B'+str(900+i)]
    prevMLUsefulness=sheet['C'+str(900+i)]
    #print(review.value)

    rm = ReviewModifier(review.value,0)
    sampleList = rm.getReviewModifier()
    sample=sampleList[0]

    sheetWrite['A'+str(i)]=id.value
    sheetWrite['B' + str(i)] = review.value
    sheetWrite['G'+str(i)]=prevMLUsefulness.value

    clf.feedReview(review.value, 0)
    nowUsefulness = clf.getPrediction()
    sheetWrite['C'+str(i)]=nowUsefulness

    '''
    clf.feedReview(review.value,0)
    myUsefulness = clf.getPrediction()
    print("________________________"+myUsefulness)

    sheetWrite['E' + str(i)] = review.value
    sheetWrite['G' + str(i)] = usefulness.value
    sheetWrite['H' + str(i)] = myUsefulness


    if(usefulness.value=='N'):
        continue

    total=total+1
    if(usefulness.value==myUsefulness):
        matched=matched+1
    '''

modified.save('PredictedSample_4rd(300).xlsx')