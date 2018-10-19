from openpyxl import Workbook, load_workbook

from Classification.Classifier import Classifier
from Classification.ReviewModifier.ReviewModifier import ReviewModifier

book = load_workbook('Final1.xlsx')
modified = Workbook()

sheet = book.active
sheetWrite = modified.active
# book.template=False

matched = 0
total = 0
for i in range(2, 1646):
    id=sheet['A'+str(i)]
    review = sheet['B' + str(i)]
    usefulness = sheet['C' + str(i)]
    #print(review.value)

    rm = ReviewModifier(review.value, 0)
    sampleList = rm.getReviewModifier()
    sample = sampleList[0]

    sheetWrite['A' + str(i)] = id.value
    sheetWrite['U' + str(i)] = usefulness.value

    sheetWrite['B' + str(i)] = sample[0]
    sheetWrite['D' + str(i)] = sample[1]
    sheetWrite['F' + str(i)] = sample[2]
    sheetWrite['J' + str(i)] = sample[3]
    sheetWrite['M' + str(i)] = sample[4]
    sheetWrite['O' + str(i)] = sample[5]
    sheetWrite['Q' + str(i)] = sample[6]
    sheetWrite['R' + str(i)] = sample[7]
    sheetWrite['S' + str(i)] = sample[8]

    '''
    clf.feedReview(review.value,0)
    myUsefulness = clf.getPrediction()
    print("________________________"+myUsefulness)

    sheetWrite['E' + str(i)] = review.value
    sheetWrite['G' + str(i)] = usefulness.value
    sheetWrite['H' + str(i)] = myUsefulness


    if(usefulness.value=='N'):
        continue

    total=total+1
    if(usefulness.value==myUsefulness):
        matched=matched+1
    '''

modified.save('Final1Judgement.xlsx')