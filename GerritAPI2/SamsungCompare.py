import openpyxl
from openpyxl import Workbook,load_workbook

from Classification.ReviewModifier.ReviewModifier import ReviewModifier

book = load_workbook('[Labled with   queries]code_reviews_to_be_labeled.xlsx')
modified=Workbook()

from Classification.Classifier import Classifier

clf = Classifier()

sheet = book.active
sheetWrite=modified.active
#book.template=False

matched=0
total=0
for i in range(2,278):
    review=sheet['A'+str(i)]
    usefulness=sheet['B'+str(i)]
    #print(review.value)

    rm = ReviewModifier(review.value,0)
    sampleList = rm.getReviewModifier()
    sample=sampleList[0]

    sheetWrite['A'+str(i)]=i
    sheetWrite['T'+str(i)]=usefulness.value

    sheetWrite['B' + str(i)] =sample[0]
    sheetWrite['D' + str(i)] =sample[1]
    sheetWrite['F' + str(i)] =sample[2]
    sheetWrite['J' + str(i)] =sample[3]
    sheetWrite['M' + str(i)] =sample[4]
    sheetWrite['O' + str(i)]=sample[5]
    sheetWrite['Q' + str(i)] = sample[6]
    sheetWrite['R' + str(i)] = sample[7]

    '''
    clf.feedReview(review.value,0)
    myUsefulness = clf.getPrediction()
    print("________________________"+myUsefulness)

    sheetWrite['E' + str(i)] = review.value
    sheetWrite['G' + str(i)] = usefulness.value
    sheetWrite['H' + str(i)] = myUsefulness

    
    if(usefulness.value=='N'):
        continue

    total=total+1
    if(usefulness.value==myUsefulness):
        matched=matched+1
    '''


modified.save('Sample_Review_Judgement.xlsx')

'''
print("Total: "+str(total))
print("Matched: "+str(matched))
print("Total Matched "+str(100.0*(matched+0.0)/total)+"%")
'''